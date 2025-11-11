# app/services/excel_processor.py
import io
import re
import traceback
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from asteval import Interpreter

# Импорт сервисов из приложения
from app.services.geocoding_service import apply_post_processing
from app.utils.helpers import get_col_from_cell
from app.services import logging_service
# --- ИМПОРТИРУЕМ ГЛОБАЛЬНЫЙ 'socketio' ---
from app.extensions import task_statuses, db, socketio

# --- УБИРАЕМ 'create_app' ОТСЮДА ---
# (app.py его уже создал, мы его импортируем через socketio)

# --- НОВЫЙ ВЫЧИСЛИТЕЛЬ ФОРМУЛ ---
_aeval = Interpreter()


def _evaluate_formula(formula_str, source_row_idx, source_ws, warnings_list):
    # ... (эта функция остается без изменений) ...
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return formula_str
    expression = formula_str[1:].strip()
    try:
        variables = set(re.findall(r'([A-Z]+\d*\{row\}?\d*)', expression, re.IGNORECASE))
        for var in variables:
            cell_ref = var.format(row=source_row_idx)
            try:
                cell_value = source_ws[cell_ref].value
                numeric_value = float(cell_value)
                expression = re.sub(r'(?i)' + re.escape(var), str(numeric_value), expression)
            except (ValueError, TypeError, AttributeError, TypeError):
                error_msg = f"Ошибка в формуле (ячейка {cell_ref}): Не удалось получить число (значение: '{cell_value}')"
                print(f"Ошибка в _evaluate_formula: {error_msg}")
                if warnings_list is not None:
                    warnings_list.append(error_msg)
                return f'#VALUE! (ссылка: {cell_ref})'
        _aeval.eval(expression)
        if _aeval.error:
            error_msg = _aeval.error_msg
            _aeval.error = None
            print(f"Ошибка asteval: {error_msg}")
            if warnings_list is not None:
                warnings_list.append(f"Ошибка вычисления ({formula_str[1:]}): {error_msg}")
            return '#NUM!'
        return _aeval.result
    except Exception as e:
        print(f"Критическая ошибка в _evaluate_formula: {e}")
        return '#ERROR!'


# --- Функции парсинга (без изменений) ---
def get_sheet_settings_map(sheet_settings):
    # ... (без изменений) ...
    settings_map = {}
    for setting in sheet_settings:
        sheet_name = setting.get('sheet_name')
        start_cell = setting.get('start_cell')
        if sheet_name and start_cell:
            start_row = int("".join(filter(str.isdigit, start_cell)))
            settings_map[sheet_name] = start_row
    return settings_map


# --- Функции применения правил (без изменений) ---
def _apply_static_value_rules(template_wb, static_value_rules, t_start_row, task_id):
    # ... (без изменений) ...
    if not static_value_rules: return
    rules_by_sheet = defaultdict(list)
    for rule in static_value_rules:
        rules_by_sheet[rule.get('target_sheet', template_wb.sheetnames[0])].append(rule)
    for sheet_name, sheet_rules in rules_by_sheet.items():
        try:
            ws = template_wb[sheet_name]
            max_row = ws.max_row
            if max_row < t_start_row + 1: continue
            for rule in sheet_rules:
                t_col_idx = column_index_from_string(rule['target_col'])
                value_to_insert = rule['value']
                for row_idx in range(t_start_row + 1, max_row + 1):
                    ws.cell(row=row_idx, column=t_col_idx).value = value_to_insert
        except KeyError:
            print(f"[{task_id}] ВНИМАНИЕ: Лист '{sheet_name}' для статичного значения не найден.")
        except Exception as e:
            print(f"[{task_id}] ОШИБКА: Ошибка применения статичного значения: {e}")


def _apply_formula_rules(source_wb, template_wb, formula_rules, sheet_settings_map, t_start_row, task_id,
                         warnings_list):
    # ... (без изменений) ...
    if not formula_rules: return
    rules_by_target_sheet = defaultdict(list)
    for rule in formula_rules:
        rules_by_target_sheet[rule.get('target_sheet', template_wb.sheetnames[0])].append(rule)
    for target_sheet_name, sheet_rules in rules_by_target_sheet.items():
        try:
            template_ws = template_wb[target_sheet_name]
            max_row = template_ws.max_row
            if max_row < t_start_row + 1: continue
            for t_row_idx in range(t_start_row + 1, max_row + 1):
                for rule in sheet_rules:
                    source_sheet_name = rule['source_sheet']
                    s_start_row = sheet_settings_map.get(source_sheet_name)
                    if s_start_row is None: continue
                    source_ws = source_wb[source_sheet_name]
                    source_row_idx = s_start_row + (t_row_idx - (t_start_row + 1))
                    formula_template = rule['formula']
                    t_col_idx = column_index_from_string(rule['target_col'])
                    calculated_value = _evaluate_formula(formula_template, source_row_idx, source_ws, warnings_list)
                    template_ws.cell(row=t_row_idx, column=t_col_idx).value = calculated_value
        except KeyError as e:
            print(f"[{task_id}] ВНИМАНИЕ: Лист '{e.args[0]}' не найден при обработке формул.")
        except Exception as e:
            print(f"[{task_id}] ОШИБКА: Ошибка применения формулы: {e}")


def _apply_cell_mappings(source_wb, template_ws, cell_mappings, task_id):
    # ... (без изменений) ...
    if not cell_mappings: return
    mappings_by_sheet = defaultdict(list)
    for mapping in cell_mappings:
        sheet_name = mapping.get('source_sheet', source_wb.sheetnames[0])
        mappings_by_sheet[sheet_name].append(mapping)
    for sheet_name, sheet_mappings in mappings_by_sheet.items():
        try:
            source_ws = source_wb[sheet_name]
        except KeyError:
            print(f"[{task_id}] ВНИМАНИЕ: Лист '{sheet_name}' для копирования ячеек не найден.")
            continue
        for mapping in sheet_mappings:
            try:
                source_cell = source_ws[mapping['source_cell']]
                dest_cell = template_ws[mapping['dest_cell']]
                dest_cell.value = source_cell.value
                if source_cell.hyperlink:
                    dest_cell.hyperlink = source_cell.hyperlink.target
                    dest_cell.style = "Hyperlink"
            except Exception as e:
                print(
                    f"[{task_id}] ОШИБКА: Ошибка при копировании ячейки {mapping['source_cell']} -> {mapping['dest_cell']}: {e}")


def _apply_source_cell_fill_rules(source_wb, template_wb, source_cell_fill_rules, t_start_row, task_id):
    # ... (без изменений) ...
    if not source_cell_fill_rules: return
    rules_by_source_sheet = defaultdict(list)
    for rule in source_cell_fill_rules:
        rules_by_source_sheet[rule.get('source_sheet', source_wb.sheetnames[0])].append(rule)
    for source_sheet_name, sheet_rules in rules_by_source_sheet.items():
        try:
            source_ws = source_wb[source_sheet_name]
        except KeyError:
            print(
                f"[{task_id}] ВНИМАНИЕ: Лист источника '{source_sheet_name}' для правила 'Заполнение из ячейки' не найден.")
            continue
        for rule in sheet_rules:
            try:
                source_cell_coord = rule['source_cell']
                value_to_insert = source_ws[source_cell_coord].value
                target_sheet_name = rule.get('target_sheet', template_wb.sheetnames[0])
                target_col = rule['target_col']
                template_ws = template_wb[target_sheet_name]
                t_col_idx = column_index_from_string(target_col)
                max_row = template_ws.max_row
                if max_row < t_start_row + 1: continue
                for row_idx in range(t_start_row + 1, max_row + 1):
                    template_ws.cell(row=row_idx, column=t_col_idx).value = value_to_insert
            except KeyError:
                print(
                    f"[{task_id}] ОШИБКА: Не найдена ячейка '{source_cell_coord}' (источник) или лист '{target_sheet_name}' (шаблон).")
            except Exception as e:
                print(f"[{task_id}] ОШИБКА: Ошибка применения правила 'Заполнение из ячейки': {e}")


def _apply_manual_rules(source_ws, template_ws, rules, s_start_row, t_start_row, used_source_cols, used_template_cols,
                        visible_rows_only, task_id,
                        sheet_name, sheet_base_progress, sheet_progress_weight):
    # ... (эта функция остается без изменений) ...
    s_end_row = source_ws.max_row
    total_rows = s_end_row - s_start_row
    if total_rows <= 0:
        print(
            f"[{task_id}] DEBUG: Лист '{sheet_name}' не содержит строк данных (s_start_row: {s_start_row}, s_end_row: {s_end_row}).")
        return
    report_interval = max(200, total_rows // 20)
    total_rules = len(rules)
    progress_weight_per_rule = (sheet_progress_weight / total_rules) if total_rules > 0 else 0
    for i, rule in enumerate(rules):
        s_col_letter = rule.get('s_col') or get_col_from_cell(rule.get('source_cell'))
        t_col_letter = rule.get('t_col') or rule.get('template_col')
        if not s_col_letter or not t_col_letter:
            continue
        s_col_idx, t_col_idx = column_index_from_string(s_col_letter), column_index_from_string(t_col_letter)
        if s_col_idx in used_source_cols or t_col_idx in used_template_cols:
            print(f"[{task_id}] DEBUG: ПРАВИЛО ПРОПУЩЕНО: Колонка {s_col_letter} или {t_col_letter} уже используется.")
            continue
        rule_base_progress = int(sheet_base_progress + (i * progress_weight_per_rule))
        target_row_counter = 0
        next_report_at = report_interval
        for r_idx in range(s_start_row + 1, s_end_row + 1):
            if visible_rows_only and source_ws.row_dimensions[r_idx].hidden:
                continue
            source_cell = source_ws.cell(row=r_idx, column=s_col_idx)
            target_cell = template_ws.cell(row=t_start_row + 1 + target_row_counter, column=t_col_idx)
            target_cell.value = source_cell.value
            if source_cell.hyperlink:
                target_cell.hyperlink = source_cell.hyperlink.target
                target_cell.style = "Hyperlink"
            target_row_counter += 1
            rows_processed = r_idx - s_start_row
            if rows_processed >= next_report_at:
                sheet_completion_ratio = rows_processed / total_rows
                total_progress = int(rule_base_progress + (sheet_completion_ratio * progress_weight_per_rule))
                _emit_status(task_id,
                             f"Лист '{sheet_name}': {rows_processed}/{total_rows} (Колонка {s_col_letter} \u2192 {t_col_letter})",
                             total_progress)
                next_report_at += report_interval
        used_source_cols.add(s_col_idx)
        used_template_cols.add(t_col_idx)
    _emit_status(task_id, f"Лист '{sheet_name}' завершен.", int(sheet_base_progress + sheet_progress_weight))


# --- Функция SocketIO (без изменений, использует глобальный socketio) ---
def _emit_status(task_id, status, progress, is_complete=False, result_ready=False, warnings=None):
    print(f"--- DEBUG [processor.py]: {task_id} - Вызов _emit_status (Progress: {progress}%) ---")

    if task_id in task_statuses:
        task_data = task_statuses[task_id]
        if task_data:
            task_data['status'] = status
            task_data['progress'] = progress

    payload = {
        'task_id': task_id,
        'status': status,
        'progress': progress,
        'result_ready': result_ready
    }

    if is_complete:
        payload['warnings'] = warnings or []

    event = 'task_complete' if is_complete else 'status_update'

    try:
        # Используем ГЛОБАЛЬНЫЙ 'socketio' (из app.extensions)
        socketio.emit(event, payload, room=task_id)
        print(f"--- DEBUG [processor.py]: {task_id} - socketio.emit УСПЕХ ---")
    except Exception as e:
        print(f"--- DEBUG [processor.py]: {task_id} - ОШИБКА при вызове socketio.emit: {e} ---")
        traceback.print_exc()


# --- Основная функция (КЛЮЧЕВОЕ ИЗМЕНЕНИЕ) ---
def process_excel_hybrid(task_id, source_file_obj, template_file_obj,
                         ranges, sheet_settings, template_rules, post_function,
                         original_template_filename, task_statuses, cell_mappings=None,
                         formula_rules=None, static_value_rules=None, visible_rows_only=False,
                         source_cell_fill_rules=None):
    print(f"--- DEBUG [processor.py]: ЗАПУСК ЗАДАЧИ {task_id} ---")

    owner_id = task_statuses.get(task_id, {}).get('owner_id')
    final_status = "Неизвестная ошибка"

    # --- УДАЛЯЕМ СОЗДАНИЕ КОНТЕКСТА ---
    # app = create_app()
    # context = app.app_context()
    # context.push()
    # --- КОНЕЦ УДАЛЕНИЯ ---

    print(f"--- DEBUG [processor.py]: {task_id} - Контекст УЖЕ должен быть (из start_background_task) ---")

    task_warnings = []

    try:
        print(f"--- DEBUG [processor.py]: {task_id} - Вход в блок TRY ---")

        _emit_status(task_id, 'Подготовка...', 5)

        print(f"--- DEBUG [processor.py]: {task_id} - _emit_status(5%) ---")

        source_wb = load_workbook(filename=source_file_obj, data_only=True)

        print(f"--- DEBUG [processor.py]: {task_id} - Source WB загружен ---")

        is_macro_enabled = original_template_filename.lower().endswith('.xlsm')
        template_wb = load_workbook(filename=template_file_obj, keep_vba=is_macro_enabled)
        template_ws = template_wb.active

        print(f"--- DEBUG [processor.py]: {task_id} - Template WB загружен ---")

        # ... (остальная логика try...catch...finally остается БЕЗ ИЗМЕНЕНИЙ) ...
        # (включая _apply_cell_mappings, _apply_manual_rules, logging_service.log_task и т.д.)

        sheet_settings_map = get_sheet_settings_map(sheet_settings)
        t_start_row = ranges.get('t_start_row', 1)
        used_template_cols = set()
        used_source_cols_by_sheet = defaultdict(set)

        # 1. Точечное копирование ячеек
        _emit_status(task_id, 'Копирую отдельные ячейки...', 10)
        _apply_cell_mappings(source_wb, template_ws, cell_mappings, task_id)

        # 1.5. Заполнение столбцов из ячейки
        _emit_status(task_id, 'Заполняю столбцы из ячеек...', 15)
        _apply_source_cell_fill_rules(source_wb, template_wb, source_cell_fill_rules, t_start_row, task_id)

        # 2. Копирование колонок
        base_progress = 20
        total_progress_weight = 50
        sheets_with_rules = set(r.get('source_sheet', source_wb.sheetnames[0]) for r in template_rules)
        sheets_to_process = [s for s in source_wb.sheetnames if s in sheets_with_rules]
        if not sheets_to_process and any(r.get('source_sheet') is None for r in template_rules):
            if source_wb.sheetnames and source_wb.sheetnames[0] not in sheets_to_process:
                sheets_to_process.append(source_wb.sheetnames[0])
        total_sheets = len(sheets_to_process)
        progress_weight_per_sheet = total_progress_weight / total_sheets if total_sheets > 0 else 0
        _emit_status(task_id, f"Найдено {total_sheets} листов для обработки колонок...", base_progress)
        for i, sheet_name in enumerate(sheets_to_process):
            try:
                source_ws = source_wb[sheet_name]
                s_start_row = sheet_settings_map.get(sheet_name, 1)
                used_source_cols = used_source_cols_by_sheet[sheet_name]
                current_template_rules = [r for r in template_rules if
                                          r.get('source_sheet', source_wb.sheetnames[0]) == sheet_name]
                if not current_template_rules:
                    continue
                sheet_base_progress = int(base_progress + (i * progress_weight_per_sheet))
                _apply_manual_rules(
                    source_ws, template_ws, current_template_rules, s_start_row, t_start_row,
                    used_source_cols,
                    used_template_cols, visible_rows_only, task_id,
                    sheet_name,
                    sheet_base_progress,
                    int(progress_weight_per_sheet)
                )
            except KeyError:
                print(f"[{task_id}] ВНИМАНИЕ: Лист '{sheet_name}' (из правил) не найден в файле-источнике.")
            except Exception as e:
                print(f"[{task_id}] ОШИБКА: Ошибка при обработке ручных правил для листа '{sheet_name}': {e}")

        # 3. Заполнение статичных значений
        _emit_status(task_id, 'Заполняю статичные значения...', 70)
        _apply_static_value_rules(template_wb, static_value_rules, t_start_row, task_id)

        # 4. Вычисление и вставка результатов формул
        _emit_status(task_id, 'Вычисляю формулы...', 80)
        _apply_formula_rules(source_wb, template_wb, formula_rules, sheet_settings_map, t_start_row, task_id,
                             task_warnings)

        # 5. Финальная пост-обработка
        _emit_status(task_id, 'Пост-обработка...', 90)
        apply_post_processing(task_id, template_wb, t_start_row, post_function, task_statuses)

        # 6. Сохранение результата
        _emit_status(task_id, 'Сохраняю результат...', 95)
        processed_file_obj = io.BytesIO()
        template_wb.save(processed_file_obj)
        processed_file_obj.seek(0)
        source_wb.close()
        template_wb.close()

        print(f"--- DEBUG [processor.py]: {task_id} - Блок TRY УСПЕШНО ЗАВЕРШЕН ---")

        # 7. Логгирование и обновление статуса (УСПЕХ)
        final_status = 'Готово!'
        # logging_service.log_task вызывается ВНУТРИ app context'а
        logging_service.log_task(
            task_id, owner_id, final_status, original_template_filename
        )
        task_statuses[task_id].update({
            'status': final_status,
            'result_file': processed_file_obj,
            'template_filename': original_template_filename,
            'warnings': task_warnings
        })
        _emit_status(task_id, final_status, 100, is_complete=True, result_ready=True, warnings=task_warnings)

    except Exception as e:
        # 8. Логгирование и обновление статуса (ОШИБКА)
        print(f"[{task_id}] КРИТИЧЕСКАЯ ОШИБКА в фоновом потоке: {e}")
        traceback.print_exc()
        final_status = f"Ошибка: {e}"
        # logging_service.log_task вызывается ВНУТРИ app context'а
        logging_service.log_task(
            task_id, owner_id, final_status, original_template_filename
        )
        task_statuses[task_id].update({
            'status': final_status,
            'result_file': None,
            'warnings': task_warnings
        })
        _emit_status(task_id, final_status, 100, is_complete=True, result_ready=False, warnings=task_warnings)

    finally:
        print(f"--- DEBUG [processor.py]: {task_id} - Вход в блок FINALLY ---")

        # --- УДАЛЯЕМ context.pop() ---
        # context.pop()
        # --- КОНЕЦ УДАЛЕНИЯ ---

        if task_id in task_statuses:
            task_data = task_statuses[task_id]
            task_statuses[task_id] = {
                'result_file': task_data.get('result_file'),
                'template_filename': task_data.get('template_filename'),
                'owner_id': task_data.get('owner_id'),
                'warnings': task_data.get('warnings')
            }

        print(f"--- DEBUG [processor.py]: {task_id} - ЗАДАЧА ЗАВЕРШЕНА ---")